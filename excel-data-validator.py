# excel-data-validator.py
# Description: Validates data in an Excel sheet across all rows and columns, checking for empty cells, and ensuring data types
# and values meet specific criteria based on the column (ID, Name, Date, Amount). Reports any inconsistencies or invalid data entries.

import openpyxl
from openpyxl import load_workbook
from datetime import datetime


def is_date(value):
    """Determines if the given value is a valid date in the 'YYYY-MM-DD' format or a datetime object."""
    if isinstance(value, datetime):
        return True
    elif isinstance(value, str):
        try:
            datetime.strptime(value, "%Y-%m-%d")
            return True
        except ValueError:
            return False
    return False


# Load the workbook and access the specified sheet
file = "path/to/your/sample_data.xlsx"  # Replace with the path to your spreadsheet
workbook = load_workbook(file)
sheet = workbook['Sample Data']  # Ensure the sheet name matches your document

# Iterate through the rows and columns to validate data
for row in sheet.iter_rows(min_row=2, values_only=False):  # Start from row 2 to skip header
    for cell in row:
        if cell.value in [None, '']:
            print(f'Cell {cell.coordinate} is empty')
            continue  # Skip further checks for empty cells

        # Validate ID: must be an integer greater than 0
        if cell.column_letter == 'A' and not (isinstance(cell.value, int) and cell.value > 0):
            print(f'Invalid ID at Cell {cell.coordinate}: {cell.value}')

        # Validate Name: must be a non-empty string
        elif cell.column_letter == 'B' and not (isinstance(cell.value, str) and cell.value.strip()):
            print(f'Invalid Name at Cell {cell.coordinate}: {cell.value}')

        # Validate Date: must be a valid date
        elif cell.column_letter == 'C' and not is_date(cell.value):
            print(f'Invalid Date at Cell {cell.coordinate}: {cell.value}')

        # Validate Amount: must be a positive float or integer
        elif cell.column_letter == 'D':
            try:
                value = float(cell.value)
                if value <= 0:
                    print(f'Invalid Amount at Cell {cell.coordinate}: {cell.value}')
            except ValueError:
                print(f'Invalid data at Cell {cell.coordinate}: {cell.value}')
