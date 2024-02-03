# excel-phone-number-standardizer.py
# Description: Searches for phone numbers in the first column of a specified Excel workbook
# and standardizes their format to "(area code) first three digits-last four digits".
# Saves the updated workbook with cleaned phone numbers.

import openpyxl
import re

# Compile a regex pattern to match phone numbers with optional country codes and various delimiters
phone_pattern = re.compile(r"(?:\+?\d{1,3}[-. ]?)?(\d{3})[-. ]?(\d{3})[-. ]?(\d{4})")
# Define the standard format for phone numbers
standard_format = "({area_code}) {first_three}-{last_four}"

# Load the workbook and select the specified sheet
wb = openpyxl.load_workbook('example_phone_numbers.xlsx')  # Replace with your actual Excel file name
sheet = wb['Sheet']  # Ensure the sheet name matches your document

# Iterate through each cell in the first column
for row in sheet.iter_rows(min_row=1, min_col=1, max_row=sheet.max_row, max_col=1):
    for cell in row:
        if isinstance(cell.value, str):  # Check if the cell contains text
            match = phone_pattern.search(cell.value)
            if match:
                # Extract parts of the phone number
                area_code, first_three, last_four = match.groups()
                # Format the number to the standard format
                standardized_number = standard_format.format(area_code=area_code, first_three=first_three, last_four=last_four)
                # Update the cell value
                cell.value = standardized_number

# Save the workbook with cleaned phone numbers under a new file name
wb.save('phone_numbers_cleaned.xlsx')
print("Phone numbers have been standardized and saved in 'phone_numbers_cleaned.xlsx'.")
