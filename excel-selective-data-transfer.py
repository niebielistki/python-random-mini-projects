# excel-selective-data-transfer.py
# Description: Identifies and transfers specified data from a "Source" sheet to a "Destination" sheet within the same Excel workbook.
# Designed to work with product names and their associated quantity and price, updating the destination sheet with selected items.

import openpyxl

# Load the workbook and select source and destination sheets
wb = openpyxl.load_workbook('sample_data_transfer.xlsx')  # Ensure this matches your file's name
source = wb['Source']  # Source sheet from where data is to be copied
dest = wb['Destination']  # Destination sheet to where data is to be pasted

# List of items to be transferred
data_to_transfer = ['Apples', 'Bananas']

# Find the next empty row in the destination sheet to start data transfer
next_empty_row = 2  # Assuming data starts from row 2

# Iterate through the source sheet to find and transfer data
for rowNum in range(2, source.max_row + 1):  # Start from row 2 to skip header
    produceName = source.cell(row=rowNum, column=1).value  # Get the name of the produce
    if produceName in data_to_transfer:
        # Retrieve quantity and price for the matching produce
        quantity_column = source.cell(row=rowNum, column=2).value
        price_column = source.cell(row=rowNum, column=3).value

        # Transfer the data to the destination sheet
        dest.cell(row=next_empty_row, column=1, value=produceName)
        dest.cell(row=next_empty_row, column=2, value=quantity_column)
        dest.cell(row=next_empty_row, column=3, value=price_column)

        # Move to the next row in the destination sheet for the next item
        next_empty_row += 1

# Save the workbook with updated data
wb.save('updated_sample_data_transfer.xlsx')
print("Data transfer complete. The updated workbook is saved as 'updated_sample_data_transfer.xlsx'.")
