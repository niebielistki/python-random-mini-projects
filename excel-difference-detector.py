# excel-difference-detector.py
# Description: Compares two Excel sheets from different workbooks and reports any differences found between them.
# It checks each cell within the range of rows and columns present in both sheets and prints discrepancies.

import openpyxl
from openpyxl import load_workbook

# Load the workbooks (Replace the paths with your actual Excel files' paths)
file1 = "path/to/your/sample_data.xlsx"
file2 = "path/to/your/sample_data_modified.xlsx"

workbook1 = load_workbook(file1)
workbook2 = load_workbook(file2)

# Access the 'Sample Data' sheet from both workbooks
sheet1 = workbook1['Sample Data']
sheet2 = workbook2['Sample Data']

# Determine the maximum row and column numbers to compare
max_row_to_compare = min(sheet1.max_row, sheet2.max_row)
max_column_to_compare = min(sheet1.max_column, sheet2.max_column)

# Iterate through each cell in the overlapping range
for row in range(1, max_row_to_compare + 1):
    for col in range(1, max_column_to_compare + 1):
        # Retrieve cell values from both sheets
        cell_value_sheet1 = sheet1.cell(row=row, column=col).value
        cell_value_sheet2 = sheet2.cell(row=row, column=col).value

        # Report differences found in cell values
        if cell_value_sheet1 != cell_value_sheet2:
            print(f"Difference at Row {row}, Column {col}: Sheet1 has '{cell_value_sheet1}', Sheet2 has '{cell_value_sheet2}'")
