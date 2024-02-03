# excel-sheets-to-csv.py
# Description: Converts each sheet in Excel files within a specified directory to individual CSV files,
# preserving data integrity.
# Usage: Perfect for users needing to automate the conversion of multiple Excel sheets to CSV format
# for data analysis or migration tasks.

import openpyxl
import os
import csv

# Directory containing the Excel files (replace with your actual directory path)
directory_path = ""  # Replace with your folder path

# List all Excel (.xlsx) files in the directory
excel_files = [file for file in os.listdir(directory_path) if file.endswith('.xlsx')]

# Process each Excel file in the list
for excel_file in excel_files:
    # Construct the full path to the Excel file
    full_path = os.path.join(directory_path, excel_file)
    wb = openpyxl.load_workbook(full_path)
    sheet_names = wb.sheetnames

    # Convert each sheet to a CSV file
    for sheet_name in sheet_names:
        csv_filename = f"{excel_file.split('.')[0]}_{sheet_name}.csv"
        csv_full_path = os.path.join(directory_path, csv_filename)  # Full path for the CSV file

        # Create and write to the CSV file
        with open(csv_full_path, 'w', newline='') as csv_file:
            writer = csv.writer(csv_file)
            sheet = wb[sheet_name]  # Access the current sheet

            # Write each row from the Excel sheet to the CSV file
            for row in sheet.iter_rows(values_only=True):
                writer.writerow(row)
