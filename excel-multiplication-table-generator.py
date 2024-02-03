# excel-multiplication-table-generator.py
# Description: Generates an Excel spreadsheet with a multiplication table of size N.
# The table size (N) is provided as a command-line argument.

import sys
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Verify command-line argument for table size
if len(sys.argv) > 1:
    try:
        N = int(sys.argv[1])
    except ValueError:
        print("Error: Please provide an integer for the multiplication table.")
        sys.exit("Usage: python multiplication-table-generator.py N")
else:
    sys.exit("Error: No size provided for the table. Usage: python multiplication-table-generator.py N")

# Create a new workbook and select the active sheet
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Multiplication Table"

# Apply bold font for the header row and column
bold_font = Font(bold=True)

# Fill in the header row and column
for i in range(1, N + 1):
    sheet.cell(row=1, column=i + 1).value = i
    sheet.cell(row=1, column=i + 1).font = bold_font
    sheet.cell(row=i + 1, column=1).value = i
    sheet.cell(row=i + 1, column=1).font = bold_font

# Populate the table with multiplication results
for row in range(2, N + 2):
    for col in range(2, N + 2):
        sheet.cell(row=row, column=col).value = (row - 1) * (col - 1)

# Adjust column width and row height for better readability
for i in range(2, N + 2):
    sheet.column_dimensions[get_column_letter(i)].width = 10
    sheet.row_dimensions[i].height = 20

# Save the workbook
wb.save('multiplication_table.xlsx')
print(f"Multiplication table for {N} has been successfully created in 'multiplication_table.xlsx'.")
