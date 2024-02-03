# excel-cell-inverter.py
# Description: Reads data from an Excel sheet and inverts the rows and columns,
# effectively transposing the cell data. The inverted data is saved in a new sheet within the same workbook.

import openpyxl
from openpyxl import load_workbook

# Load the workbook and select the active sheet
wb = openpyxl.load_workbook('example_cell_inverter.xlsx')  # Replace with your actual Excel file name
sheet = wb.active  # Assumes data to be inverted is in the active sheet

# Container to hold the original sheet's data by columns
hold_cell_data = []

# Extract data from each column in the active sheet
for col in range(1, sheet.max_column + 1):
    column_data = [sheet.cell(row=row, column=col).value for row in range(1, sheet.max_row + 1)]
    hold_cell_data.append(column_data)

# Create a new sheet for the inverted data
inv_sheet = wb.create_sheet(title='Inverted Sheet')

# Populate the new sheet with inverted data
for col_idx, col_data in enumerate(hold_cell_data, start=1):
    for row_idx, cell_value in enumerate(col_data, start=1):
        inv_sheet.cell(row=col_idx, column=row_idx, value=cell_value)

# Save the workbook with the new inverted data sheet
wb.save('inverted_example_cell_inverter.xlsx')
print("Inverted sheet has been added and saved to 'inverted_example_cell_inverter.xlsx'.")
