# excel-blank-row-inserter.py
# Description: Inserts a specified number of blank rows into an existing Excel workbook starting from a given row.
# The modified workbook is saved under a new filename.

import sys
import openpyxl

# Function to insert blank rows and save the workbook with a new filename
def insert_blank_rows(start_row, num_blank_rows, source_filename, updated_filename):
    wb = openpyxl.load_workbook(source_filename)
    sheet = wb.active

    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active

    # Copy rows from the original sheet to the new sheet up to the start row
    for row_num, row in enumerate(sheet.iter_rows(min_row=1, max_row=start_row - 1), start=1):
        for cell in row:
            new_sheet.cell(row=row_num, column=cell.column).value = cell.value

    # Offset for new rows
    new_row_num = start_row + num_blank_rows

    # Copy remaining rows from the original sheet to the new sheet starting from the offset
    for row in sheet.iter_rows(min_row=start_row):
        for cell in row:
            new_sheet.cell(row=new_row_num, column=cell.column).value = cell.value
        new_row_num += 1

    new_wb.save(updated_filename)
    print(f"New file saved as: {updated_filename}")

# Path to the source Excel file (replace with your actual file path)
source_filename = "path/to/your/file.xlsx"

if len(sys.argv) == 4:
    try:
        N = int(sys.argv[1])
        M = int(sys.argv[2])
        updated_filename = sys.argv[3]

        if not updated_filename.endswith('.xlsx'):
            print("The file name does not indicate an Excel workbook. Please provide a .xlsx file.")
        else:
            insert_blank_rows(N, M, source_filename, updated_filename)
    except ValueError:
        print("The N and M arguments must be integers. Please provide correct values.")
else:
    print("Incorrect number of arguments. Please provide N, M, and the filename.")
