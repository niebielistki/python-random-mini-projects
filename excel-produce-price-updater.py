# excel-produce-price-updater.py
# Description: Scans a produce sales spreadsheet for specific items and updates their prices according to a predefined list.
# This ensures pricing information is current and accurate across the sales data.

import openpyxl

# Load the workbook and access the specified sheet
wb = openpyxl.load_workbook('produceSales.xlsx')  # Replace with your actual Excel file name
sheet = wb['Sheet']  # Ensure the sheet name matches your document

# Dictionary of produce items and their updated prices
PRICE_UPDATES = {
    'Garlic': 3.07,
    'Celery': 1.19,
    'Lemon': 1.27
} # You can also updated it with your name of products and updated prices

# Iterate over the rows to find and update the prices of specified produce items
for rowNum in range(2, sheet.max_row + 1):  # Start from row 2 to skip the header
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        # Update the price in the spreadsheet
        sheet.cell(row=rowNum, column=2, value=PRICE_UPDATES[produceName])

# Save the workbook with updated prices
wb.save('updatedProduceSales.xlsx')
print("Produce prices updated successfully in 'updatedProduceSales.xlsx'.")
