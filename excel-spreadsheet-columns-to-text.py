# excel-spreadsheet-columns-to-text.py
# Description: Extracts text from each column of a specified Excel spreadsheet and saves the content of each column into separate text files. It skips empty columns and names files sequentially.

import openpyxl

# Function to check if a column in the sheet is empty
def is_column_empty(sheet, column_index):
    return all(cell[0].value is None for cell in sheet.iter_rows(min_col=column_index, max_col=column_index))

# Function to create a new text file from provided text content
def create_new_file(file_path, text):
    with open(file_path, 'w') as file:
        file.write(text)

# Path to the Excel file (replace with your actual file path)
excel_file = ""

# Load the workbook and select the active sheet
wb = openpyxl.load_workbook(excel_file)
sheet = wb.active  # Assuming the relevant data is in the active sheet

# Determine the number of columns with data
columns_with_data = sum(not is_column_empty(sheet, col) for col in range(1, sheet.max_column + 1))

# Iterate through each column with data and create a corresponding text file
for col in range(1, columns_with_data + 1):
    text_list = [str(cell[0].value) for cell in sheet.iter_rows(min_col=col, max_col=col, values_only=True) if cell[0] is not None]
    column_filename = f"column{col}.txt"
    create_new_file(column_filename, '\n'.join(text_list))
    print(f"The file {column_filename} was created.")

print('The text from all columns was saved in the corresponding files.')
